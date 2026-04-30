from __future__ import annotations

import asyncio
import csv
import json
from pathlib import Path
from typing import Any

import openpyxl

from .change_proposal import ChangeOp, ChangeProposal
from .config import ProjectConfig
from .models import TableIndex
from .paths import get_tables_dir
from .table_indexer import TableIndexer


class ApplyError(Exception):
    def __init__(self, op: ChangeOp, reason: str):
        self.op = op
        self.reason = reason
        super().__init__(f"{op.op} on {op.table}#{op.row_id}: {reason}")


# Suffixes treated as table sources we can edit.
_TABLE_SUFFIXES = {".csv", ".xlsx", ".xls", ".txt"}


class ChangeApplier:
    def __init__(
        self,
        project: ProjectConfig,
        svn_root: Path,
        table_indexer: TableIndexer | None = None,
    ):
        self.project = project
        self.svn_root = Path(svn_root)
        self.table_indexer = table_indexer

    async def dry_run(self, proposal: ChangeProposal) -> list[dict]:
        return await asyncio.to_thread(self._dry_run_sync, proposal)

    async def apply(self, proposal: ChangeProposal) -> dict:
        return await asyncio.to_thread(self._apply_sync, proposal)

    def read_rows(
        self, table_name: str, offset: int = 0, limit: int = 100
    ) -> dict[str, Any]:
        """Return paginated raw rows of a table for grid views.

        Output shape:
            { headers, rows, total, header_row, comment_row, source }
        """
        source = self._resolve_source_file(table_name)
        workbook = self._load_table(source)
        rows = workbook["rows"]
        header_row = self.project.table_convention.header_row
        comment_row = getattr(self.project.table_convention, "comment_row", 0) or 0
        header_idx = max(header_row - 1, 0)
        if len(rows) <= header_idx:
            headers: list[str] = []
            data_start = 0
        else:
            headers = [
                "" if v is None else str(v).strip() for v in rows[header_idx]
            ]
            # data starts after header row and (optional) comment row
            data_start = header_idx + 1
            if comment_row and comment_row > header_row:
                data_start = max(data_start, comment_row)
        # de-duplicate empty headers to stable column keys
        out_headers: list[str] = []
        seen: dict[str, int] = {}
        for i, h in enumerate(headers):
            key = h or f"Column_{i}"
            if key in seen:
                seen[key] += 1
                key = f"{key}_{seen[key]}"
            else:
                seen[key] = 0
            out_headers.append(key)

        body = rows[data_start:] if data_start < len(rows) else []
        # drop fully-empty trailing rows
        while body and all(v is None or str(v).strip() == "" for v in body[-1]):
            body.pop()
        total = len(body)
        sliced = body[offset : offset + limit]
        normalized = [
            [None if v is None else (v if isinstance(v, (str, int, float, bool)) else str(v))
             for v in row]
            for row in sliced
        ]
        rel_source = ""
        try:
            rel_source = str(source.relative_to(self.svn_root)).replace("\\", "/")
        except Exception:
            rel_source = str(source)
        return {
            "headers": out_headers,
            "rows": normalized,
            "total": total,
            "header_row": header_row,
            "comment_row": comment_row,
            "source": rel_source,
        }

    def _dry_run_sync(self, proposal: ChangeProposal) -> list[dict]:
        grouped = self._group_ops_by_file(proposal.ops)
        previews: list[dict] = []
        for source, ops in grouped.items():
            workbook = self._load_table(source)
            for op in ops:
                try:
                    before, after = self._apply_op(workbook, op, mutate=False)
                    previews.append(
                        {
                            "op": op.model_dump(mode="json"),
                            "before": before,
                            "after": after,
                            "ok": True,
                            "reason": None,
                        }
                    )
                except ApplyError as exc:
                    previews.append(
                        {
                            "op": op.model_dump(mode="json"),
                            "before": None,
                            "after": None,
                            "ok": False,
                            "reason": exc.reason,
                        }
                    )
        return previews

    def _apply_sync(self, proposal: ChangeProposal) -> dict:
        grouped = self._group_ops_by_file(proposal.ops)
        changed_files: list[str] = []
        counts = {"update_cell": 0, "insert_row": 0, "delete_row": 0}
        for source, ops in grouped.items():
            pending = source.with_name(f"{source.name}.ltclaw_pending")
            try:
                workbook = self._load_table(source)
                for op in ops:
                    self._apply_op(workbook, op, mutate=True)
                    counts[op.op] += 1
                self._write_pending(workbook, source, pending)
                self._load_table(pending, format_source=source)
                pending.replace(source)
                changed_files.append(str(source.relative_to(self.svn_root)).replace("\\", "/"))
            except Exception:
                if pending.exists():
                    pending.unlink()
                raise
        return {
            "changed_files": changed_files,
            "summary": (
                f"{counts['update_cell']} updates / "
                f"{counts['insert_row']} inserts / "
                f"{counts['delete_row']} deletes"
            ),
        }

    def _group_ops_by_file(self, ops: list[ChangeOp]) -> dict[Path, list[ChangeOp]]:
        grouped: dict[Path, list[ChangeOp]] = {}
        for op in ops:
            source = self._resolve_source_file(op.table)
            grouped.setdefault(source, []).append(op)
        return grouped

    def _resolve_source_file(self, table_name: str) -> Path:
        indexed = self._resolve_source_from_index(table_name)
        if indexed is not None:
            return indexed
        candidates: list[Path] = []
        for path in self.svn_root.rglob("*"):
            if not path.is_file():
                continue
            if path.suffix.lower() not in _TABLE_SUFFIXES:
                continue
            if path.stem != table_name:
                continue
            relative = path.relative_to(self.svn_root)
            if self._is_table_path(relative):
                candidates.append(path)
        if not candidates:
            raise ApplyError(
                ChangeOp(op="update_cell", table=table_name, row_id="?"),
                "source table file not found",
            )
        if len(candidates) > 1:
            raise ApplyError(
                ChangeOp(op="update_cell", table=table_name, row_id="?"),
                "multiple source table files found",
            )
        return candidates[0]

    def _resolve_source_from_index(self, table_name: str) -> Path | None:
        index_file = get_tables_dir(self.svn_root) / f"{table_name}.json"
        if not index_file.exists():
            return None
        data = json.loads(index_file.read_text(encoding="utf-8"))
        source_path = TableIndex.model_validate(data).source_path
        return self.svn_root / source_path

    def _is_table_path(self, relative_path: Path) -> bool:
        if not self.project.paths:
            return True
        path_str = str(relative_path).replace("\\", "/")
        table_rules = [rule for rule in self.project.paths if rule.semantic == "table"]
        if not table_rules:
            return True
        return any(relative_path.match(rule.path) or path_str == rule.path for rule in table_rules)

    def _load_table(self, source: Path, format_source: Path | None = None) -> dict:
        suffix = (format_source or source).suffix.lower()
        if suffix in {".xlsx", ".xls"}:
            if source.suffix.lower() == ".ltclaw_pending":
                with source.open("rb") as fh:
                    workbook = openpyxl.load_workbook(fh)
            else:
                workbook = openpyxl.load_workbook(source)
            sheet = workbook.active
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
            return {
                "kind": "xlsx",
                "source": source,
                "workbook": workbook,
                "sheet": sheet,
                "rows": rows,
            }
        if suffix == ".csv":
            encoding, delimiter, rows = self._read_csv_with_metadata(source)
            return {
                "kind": "csv",
                "source": source,
                "encoding": encoding,
                "delimiter": delimiter,
                "rows": rows,
            }
        if suffix == ".txt":
            encoding, rows = self._read_txt_with_metadata(source)
            return {
                "kind": "txt",
                "source": source,
                "encoding": encoding,
                "rows": rows,
            }
        raise ValueError(f"unsupported file type: {suffix}")

    def _read_csv_with_metadata(self, source: Path) -> tuple[str, str, list[list[Any]]]:
        encodings = ["utf-8", "gbk", "gb2312", "utf-8-sig"]
        content = None
        used_encoding = "utf-8"
        for encoding in encodings:
            try:
                content = source.read_text(encoding=encoding)
                used_encoding = encoding
                break
            except UnicodeDecodeError:
                continue
        if content is None:
            raise ValueError("unable to decode csv file")
        try:
            delimiter = csv.Sniffer().sniff(content[:1024]).delimiter
        except Exception:
            delimiter = ","
        rows = list(csv.reader(content.splitlines(), delimiter=delimiter))
        return used_encoding, delimiter, rows

    def _read_txt_with_metadata(self, source: Path) -> tuple[str, list[list[Any]]]:
        encodings = ["utf-8-sig", "utf-8", "gbk", "gb2312"]
        content = None
        used_encoding = "utf-8"
        for encoding in encodings:
            try:
                content = source.read_text(encoding=encoding)
                used_encoding = encoding
                break
            except UnicodeDecodeError:
                continue
        if content is None:
            raise ValueError("unable to decode txt file")
        # Tab-delimited rows. The header cell may carry inline type/doc
        # annotations like ``Name\u2507Type=string;Doc=...``; we strip those.
        rows: list[list[Any]] = []
        for raw_line in content.splitlines():
            cells = raw_line.split("\t")
            cleaned: list[Any] = []
            for cell in cells:
                # Drop annotation suffix on header cells; keep value otherwise.
                if "\u2507" in cell:
                    cell = cell.split("\u2507", 1)[0]
                cleaned.append(cell)
            rows.append(cleaned)
        return used_encoding, rows

    def _write_pending(self, workbook: dict, source: Path, pending: Path) -> None:
        if workbook["kind"] == "csv":
            with pending.open("w", encoding=workbook["encoding"], newline="") as fh:
                writer = csv.writer(fh, delimiter=workbook["delimiter"])
                writer.writerows(workbook["rows"])
            return
        if workbook["kind"] == "xlsx":
            workbook["workbook"].save(pending)
            return
        if workbook["kind"] == "txt":
            lines: list[str] = []
            for row in workbook["rows"]:
                cells = ["" if v is None else str(v) for v in row]
                lines.append("\t".join(cells))
            text = "\n".join(lines)
            if not text.endswith("\n"):
                text += "\n"
            pending.write_text(text, encoding=workbook["encoding"])
            return
        raise ValueError(f"unsupported workbook kind: {workbook['kind']}")

    def _apply_op(self, workbook: dict, op: ChangeOp, mutate: bool) -> tuple[Any, Any]:
        rows = workbook["rows"]
        header_idx = self.project.table_convention.header_row - 1
        if len(rows) <= header_idx:
            raise ApplyError(op, "header row not found")
        headers = ["" if value is None else str(value).strip() for value in rows[header_idx]]
        # Case-insensitive field map (with case-sensitive priority).
        field_map: dict[str, int] = {}
        field_map_ci: dict[str, int] = {}
        for index, name in enumerate(headers):
            if not name:
                continue
            field_map[name] = index
            field_map_ci.setdefault(name.lower(), index)
        primary_key = self.project.table_convention.resolve_primary_key(
            table_name=op.table,
            headers=headers,
        )
        pk_index = field_map.get(primary_key)
        if pk_index is None:
            pk_index = field_map_ci.get(primary_key.lower())
        if pk_index is None:
            raise ApplyError(op, f"primary key field not found: {primary_key}")
        data_start = header_idx + 1
        comment_row = getattr(self.project.table_convention, "comment_row", 0) or 0
        if comment_row and comment_row > self.project.table_convention.header_row:
            data_start = max(data_start, comment_row)
        row_index = self._find_row_index(rows, data_start, pk_index, op.row_id)

        def _resolve_field(name: str) -> int | None:
            if not name:
                return None
            if name in field_map:
                return field_map[name]
            return field_map_ci.get(name.lower())

        if op.op == "update_cell":
            if row_index is None:
                raise ApplyError(op, "row_id not found")
            col_index = _resolve_field(op.field or "")
            if col_index is None:
                raise ApplyError(op, f"field not found: {op.field}")
            before = rows[row_index][col_index] if col_index < len(rows[row_index]) else None
            after = self._coerce_value(
                op.new_value,
                before,
                self._column_samples(rows, data_start, col_index),
                op,
            )
            if mutate:
                self._ensure_row_len(rows[row_index], len(headers))
                rows[row_index][col_index] = after
                self._sync_row_to_sheet(workbook, row_index)
            return before, after

        if op.op == "insert_row":
            if not isinstance(op.new_value, dict):
                raise ApplyError(op, "new_value must be a dict for insert_row")
            if row_index is not None:
                raise ApplyError(op, "row_id already exists")
            row_values = [None] * len(headers)
            samples_by_col = {
                col_index: self._column_samples(rows, data_start, col_index)
                for col_index in range(len(headers))
            }
            for field_name, value in op.new_value.items():
                col_index = _resolve_field(field_name)
                if col_index is None:
                    raise ApplyError(op, f"field not found: {field_name}")
                row_values[col_index] = self._coerce_value(value, None, samples_by_col[col_index], op)
            row_values[pk_index] = self._coerce_value(op.row_id, None, samples_by_col[pk_index], op)
            before = None
            after = {headers[i]: row_values[i] for i in range(len(headers))}
            if mutate:
                rows.append(row_values)
                self._append_row_to_sheet(workbook, row_values)
            return before, after

        if op.op == "delete_row":
            if row_index is None:
                raise ApplyError(op, "row_id not found")
            before = self._row_to_dict(rows[row_index], headers)
            if mutate:
                del rows[row_index]
                self._delete_row_from_sheet(workbook, row_index)
            return before, None

        raise ApplyError(op, f"unsupported op: {op.op}")

    def _find_row_index(
        self,
        rows: list[list[Any]],
        data_start: int,
        pk_index: int,
        row_id: str | int,
    ) -> int | None:
        target = str(row_id)
        for row_index in range(data_start, len(rows)):
            row = rows[row_index]
            if pk_index >= len(row):
                continue
            cell = row[pk_index]
            if cell is None:
                continue
            if str(cell) == target:
                return row_index
        return None

    def _column_samples(
        self,
        rows: list[list[Any]],
        data_start: int,
        col_index: int,
    ) -> list[Any]:
        samples = []
        for row in rows[data_start:]:
            if col_index < len(row) and row[col_index] is not None:
                samples.append(row[col_index])
        return samples

    def _coerce_value(
        self,
        value: Any,
        existing_value: Any,
        column_samples: list[Any],
        op: ChangeOp,
    ) -> Any:
        target = existing_value
        if target is None:
            for sample in column_samples:
                if sample is not None:
                    target = sample
                    break
        inferred_type = self._infer_column_type(column_samples)
        if target is None:
            return self._coerce_by_inferred_type(value, inferred_type, op)
        try:
            if inferred_type == "int":
                return int(value)
            if inferred_type == "float":
                return float(value)
            if isinstance(target, bool):
                if isinstance(value, bool):
                    return value
                lowered = str(value).strip().lower()
                if lowered in {"true", "1", "yes"}:
                    return True
                if lowered in {"false", "0", "no"}:
                    return False
                raise ValueError("invalid bool value")
            if isinstance(target, int) and not isinstance(target, bool):
                return int(value)
            if isinstance(target, float):
                return float(value)
            if isinstance(target, str):
                return str(value)
            return value
        except Exception as exc:
            raise ApplyError(op, f"type conversion failed: {value!r}") from exc

    def _infer_column_type(self, column_samples: list[Any]) -> str | None:
        non_empty = [sample for sample in column_samples if sample is not None and str(sample).strip() != ""]
        if not non_empty:
            return None
        if all(self._is_int_like(sample) for sample in non_empty):
            return "int"
        if all(self._is_float_like(sample) for sample in non_empty):
            return "float"
        return None

    def _coerce_by_inferred_type(
        self,
        value: Any,
        inferred_type: str | None,
        op: ChangeOp,
    ) -> Any:
        try:
            if inferred_type == "int":
                return int(value)
            if inferred_type == "float":
                return float(value)
            return value
        except Exception as exc:
            raise ApplyError(op, f"type conversion failed: {value!r}") from exc

    def _is_int_like(self, value: Any) -> bool:
        if isinstance(value, bool):
            return False
        if isinstance(value, int):
            return True
        if isinstance(value, float):
            return value.is_integer()
        try:
            int(str(value))
            return True
        except Exception:
            return False

    def _is_float_like(self, value: Any) -> bool:
        if isinstance(value, bool):
            return False
        if isinstance(value, (int, float)):
            return True
        try:
            float(str(value))
            return True
        except Exception:
            return False

    def _row_to_dict(self, row: list[Any], headers: list[str]) -> dict[str, Any]:
        result: dict[str, Any] = {}
        for index, header in enumerate(headers):
            result[header] = row[index] if index < len(row) else None
        return result

    def _ensure_row_len(self, row: list[Any], length: int) -> None:
        while len(row) < length:
            row.append(None)

    def _sync_row_to_sheet(self, workbook: dict, row_index: int) -> None:
        if workbook["kind"] != "xlsx":
            return
        row = workbook["rows"][row_index]
        for col_index, value in enumerate(row, start=1):
            workbook["sheet"].cell(row=row_index + 1, column=col_index, value=value)

    def _append_row_to_sheet(self, workbook: dict, row_values: list[Any]) -> None:
        if workbook["kind"] != "xlsx":
            return
        workbook["sheet"].append(row_values)

    def _delete_row_from_sheet(self, workbook: dict, row_index: int) -> None:
        if workbook["kind"] != "xlsx":
            return
        workbook["sheet"].delete_rows(row_index + 1)
