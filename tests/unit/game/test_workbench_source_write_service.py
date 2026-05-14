from __future__ import annotations

import json
from pathlib import Path
from unittest.mock import AsyncMock, MagicMock

import openpyxl
import pytest

from ltclaw_gy_x.game.change_applier import ChangeApplier
from ltclaw_gy_x.game.config import (
    FilterConfig,
    PathRule,
    ProjectConfig,
    ProjectMeta,
    SvnConfig,
    TableConvention,
)
from ltclaw_gy_x.game.knowledge_release_store import (
    create_release,
    get_current_release,
    set_current_release,
)
from ltclaw_gy_x.game.knowledge_release_builders import DEFAULT_RELEASE_INDEXES
from ltclaw_gy_x.game.models import KnowledgeIndexArtifact, KnowledgeManifest, KnowledgeMap
from ltclaw_gy_x.game.workbench_source_write_service import (
    WorkbenchSourceWriteOp,
    WorkbenchSourceWriteService,
)


def _project_config(svn_root: Path) -> ProjectConfig:
    return ProjectConfig(
        project=ProjectMeta(name="Test Game", engine="Unity", language="zh-CN"),
        svn=SvnConfig(root=str(svn_root), poll_interval_seconds=300, jitter_seconds=30),
        paths=[PathRule(path="tables/*", semantic="table", system="core")],
        filters=FilterConfig(include_ext=[".xlsx", ".csv", ".txt"], exclude_glob=[]),
        table_convention=TableConvention(header_row=1, primary_key_field="ID"),
        doc_templates={},
        models={},
    )


def _write_csv(path: Path, rows: list[list[object]], delimiter: str = ",") -> None:
    content = "\n".join(delimiter.join("" if value is None else str(value) for value in row) for row in rows)
    path.write_text(content + "\n", encoding="utf-8")


def _read_csv(path: Path, delimiter: str = ",") -> list[list[str]]:
    return [line.split(delimiter) for line in path.read_text(encoding="utf-8").strip().splitlines()]


def _write_xlsx(path: Path, rows: list[list[object]]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    wb.save(path)


def _read_xlsx(path: Path) -> list[list[object]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    return [list(row) for row in ws.iter_rows(values_only=True)]


def _write_txt(path: Path, lines: list[str]) -> None:
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _read_text(path: Path) -> str:
    return path.read_text(encoding="utf-8")


@pytest.fixture
def sample_env(tmp_path, monkeypatch):
    monkeypatch.setenv("LTCLAW_WORKING_DIR", str(tmp_path / "working"))
    svn_root = tmp_path / "svn"
    tables_dir = svn_root / "tables"
    tables_dir.mkdir(parents=True)

    hero_csv = tables_dir / "Hero.csv"
    item_xlsx = tables_dir / "Item.xlsx"
    skill_txt = tables_dir / "Skill.txt"
    legacy_xls = tables_dir / "Legacy.xls"
    unknown_json = tables_dir / "Weird.json"

    _write_csv(hero_csv, [["ID", "Name", "HP"], [1, "Knight", 100], [2, "Mage", 80]])
    _write_xlsx(item_xlsx, [["ID", "Name", "Price"], [10, "Potion", 5], [20, "Ether", 12]])
    _write_txt(
        skill_txt,
        [
            "ID┃Type=int	Name┃Type=string	Damage┃Type=int",
            "1	Slash	100",
            "2	Fireball	120",
        ],
    )
    legacy_xls.write_text("legacy", encoding="utf-8")
    unknown_json.write_text("{}\n", encoding="utf-8")

    return {
        "svn_root": svn_root,
        "workspace_dir": tmp_path / "workspace",
        "hero_csv": hero_csv,
        "item_xlsx": item_xlsx,
        "skill_txt": skill_txt,
        "legacy_xls": legacy_xls,
        "unknown_json": unknown_json,
        "project": _project_config(svn_root),
    }


def _service(sample_env) -> WorkbenchSourceWriteService:
    applier = ChangeApplier(sample_env["project"], sample_env["svn_root"])
    return WorkbenchSourceWriteService(
        change_applier=applier,
        workspace_dir=sample_env["workspace_dir"],
        agent_id="agent-1",
        session_id="session-1",
    )


def _read_audit_lines(service: WorkbenchSourceWriteService) -> list[dict]:
    audit_path = service._audit_path()
    assert audit_path.exists()
    return [json.loads(line) for line in audit_path.read_text(encoding="utf-8").splitlines() if line.strip()]


class _JsonBoom:
    pass


@pytest.mark.asyncio
async def test_update_cell_can_write_csv_and_record_audit(sample_env):
    service = _service(sample_env)

    outcome = await service.write(
        ops=[WorkbenchSourceWriteOp(op="update_cell", table="Hero", row_id=1, field="HP", new_value=150)],
        reason="raise hp",
    )

    assert outcome.ok is True
    assert outcome.payload["svn_update_required"] is True
    assert outcome.payload["svn_update_warning"] == service.SVN_UPDATE_WARNING
    assert outcome.payload["release_id_at_write"] is None
    assert outcome.payload["source_files"] == ["tables/Hero.csv"]
    assert outcome.payload["changes"] == [
        {
            "op": "update_cell",
            "table": "Hero",
            "row_id": 1,
            "field": "HP",
            "old_value": "100",
            "new_value": 150,
        }
    ]
    assert outcome.payload["audit_recorded"] is True
    assert _read_csv(sample_env["hero_csv"])[1] == ["1", "Knight", "150"]
    audit_record = _read_audit_lines(service)[0]
    assert audit_record["event_type"] == "workbench.source.write"
    assert audit_record["agent_id"] == "agent-1"
    assert audit_record["session_id"] == "session-1"
    assert audit_record["time"]
    assert audit_record["release_id_at_write"] is None
    assert audit_record["reason"] == "raise hp"
    assert audit_record["source_files"] == ["tables/Hero.csv"]
    assert audit_record["success"] is True
    assert audit_record["changes"][0]["old_value"] == "100"
    assert audit_record["changes"][0]["new_value"] == 150


@pytest.mark.asyncio
async def test_source_write_does_not_call_svn_update_commit_or_revert(sample_env):
    service = _service(sample_env)
    service.change_applier.svn_update = AsyncMock()
    service.change_applier.svn_commit = AsyncMock()
    service.change_applier.svn_revert = AsyncMock()
    service.change_applier.update = AsyncMock()
    service.change_applier.commit = AsyncMock()
    service.change_applier.revert = AsyncMock()
    service.change_applier.trigger_now = AsyncMock()
    forbidden_committer = MagicMock()
    forbidden_committer.commit = AsyncMock()
    forbidden_committer.revert = AsyncMock()
    service.change_applier.svn_committer = forbidden_committer

    outcome = await service.write(
        ops=[WorkbenchSourceWriteOp(op="update_cell", table="Hero", row_id=1, field="HP", new_value=150)],
        reason="raise hp",
    )

    assert outcome.ok is True
    service.change_applier.svn_update.assert_not_awaited()
    service.change_applier.svn_commit.assert_not_awaited()
    service.change_applier.svn_revert.assert_not_awaited()
    service.change_applier.update.assert_not_awaited()
    service.change_applier.commit.assert_not_awaited()
    service.change_applier.revert.assert_not_awaited()
    service.change_applier.trigger_now.assert_not_awaited()
    forbidden_committer.commit.assert_not_awaited()
    forbidden_committer.revert.assert_not_awaited()


@pytest.mark.asyncio
async def test_insert_row_can_write_and_record_audit(sample_env):
    service = _service(sample_env)

    outcome = await service.write(
        ops=[
            WorkbenchSourceWriteOp(
                op="insert_row",
                table="Hero",
                row_id=3,
                new_value={"ID": 3, "Name": "Archer", "HP": 90},
            )
        ],
        reason="add archer",
    )

    assert outcome.ok is True
    assert outcome.payload["svn_update_required"] is True
    assert _read_csv(sample_env["hero_csv"])[3 - 0] == ["3", "Archer", "90"]
    audit_record = _read_audit_lines(service)[0]
    assert audit_record["changes"][0]["op"] == "insert_row"
    assert audit_record["changes"][0]["old_value"] is None
    assert audit_record["changes"][0]["new_value"]["Name"] == "Archer"


@pytest.mark.asyncio
async def test_update_cell_can_write_xlsx(sample_env):
    service = _service(sample_env)

    outcome = await service.write(
        ops=[WorkbenchSourceWriteOp(op="update_cell", table="Item", row_id=10, field="Price", new_value=8)],
        reason="discount item",
    )

    assert outcome.ok is True
    assert outcome.payload["source_files"] == ["tables/Item.xlsx"]
    assert _read_xlsx(sample_env["item_xlsx"])[1] == [10, "Potion", 8]


@pytest.mark.asyncio
async def test_txt_write_preserves_header_metadata(sample_env):
    service = _service(sample_env)

    outcome = await service.write(
        ops=[WorkbenchSourceWriteOp(op="update_cell", table="Skill", row_id=2, field="Damage", new_value=135)],
        reason="buff skill",
    )

    assert outcome.ok is True
    lines = _read_text(sample_env["skill_txt"]).splitlines()
    assert lines[0] == "ID┃Type=int	Name┃Type=string	Damage┃Type=int"
    assert lines[2] == "2	Fireball	135"


@pytest.mark.asyncio
async def test_delete_row_is_blocked_and_failure_is_audited(sample_env):
    service = _service(sample_env)

    outcome = await service.write(
        ops=[WorkbenchSourceWriteOp(op="delete_row", table="Hero", row_id=1)],
        reason="forbidden delete",
    )

    assert outcome.ok is False
    assert outcome.status_code == 400
    assert "delete_row is blocked" in outcome.payload["message"]
    assert outcome.payload["svn_update_required"] is True
    assert outcome.payload["source_files"] == []
    assert outcome.payload["changes"] == []
    assert outcome.payload["audit_recorded"] is True
    audit_record = _read_audit_lines(service)[0]
    assert audit_record["success"] is False
    assert "delete_row is blocked" in audit_record["failure"]


@pytest.mark.asyncio
@pytest.mark.parametrize(
    "op,message",
    [
        (
            WorkbenchSourceWriteOp(op="schema_update", table="Hero", row_id=1, field="HP", new_value=200),
            "Schema and unsupported ops are blocked",
        ),
        (
            WorkbenchSourceWriteOp(op="update_cell", table="Hero", row_id=1, field="ID", new_value=9),
            "Changing the primary key is blocked",
        ),
        (
            WorkbenchSourceWriteOp(op="insert_row", table="Hero", row_id=3, new_value={"ID": 3, "Name": "Archer", "MP": 10}),
            "Adding new fields is blocked",
        ),
        (
            WorkbenchSourceWriteOp(op="update_cell", table="MissingTable", row_id=1, field="HP", new_value=1),
            "source table file not found",
        ),
    ],
)
async def test_blocked_schema_new_field_new_table_and_primary_key_requests(sample_env, op, message):
    service = _service(sample_env)

    outcome = await service.write(ops=[op], reason="blocked")

    assert outcome.ok is False
    assert outcome.status_code == 400
    assert message in outcome.payload["message"]
    assert outcome.payload["audit_recorded"] is True


@pytest.mark.asyncio
async def test_xls_is_not_supported_for_source_write(sample_env):
    service = _service(sample_env)

    outcome = await service.write(
        ops=[WorkbenchSourceWriteOp(op="update_cell", table="Legacy", row_id=1, field="HP", new_value=1)],
        reason="legacy xls",
    )

    assert outcome.ok is False
    assert outcome.status_code == 400
    assert ".xls" in outcome.payload["message"]
    assert outcome.payload["audit_recorded"] is True
    audit_record = _read_audit_lines(service)[0]
    assert audit_record["success"] is False
    assert ".xls" in audit_record["failure"]


@pytest.mark.asyncio
async def test_unknown_file_format_is_blocked_and_failure_is_audited(sample_env, monkeypatch):
    service = _service(sample_env)

    monkeypatch.setattr(service.change_applier, "get_source_file", lambda _table: sample_env["unknown_json"])

    outcome = await service.write(
        ops=[WorkbenchSourceWriteOp(op="update_cell", table="Weird", row_id=1, field="HP", new_value=1)],
        reason="unknown format",
    )

    assert outcome.ok is False
    assert outcome.status_code == 400
    assert "only supports .csv, .xlsx, and .txt" in outcome.payload["message"]
    assert outcome.payload["audit_recorded"] is True
    audit_record = _read_audit_lines(service)[0]
    assert audit_record["success"] is False
    assert "only supports .csv, .xlsx, and .txt" in audit_record["failure"]


@pytest.mark.asyncio
async def test_update_cell_unknown_field_is_blocked_and_failure_is_audited(sample_env):
    service = _service(sample_env)

    outcome = await service.write(
        ops=[WorkbenchSourceWriteOp(op="update_cell", table="Hero", row_id=1, field="MP", new_value=1)],
        reason="unknown field",
    )

    assert outcome.ok is False
    assert outcome.status_code == 400
    assert "Updating unknown fields is blocked" in outcome.payload["message"]
    assert outcome.payload["audit_recorded"] is True
    audit_record = _read_audit_lines(service)[0]
    assert audit_record["success"] is False
    assert "Updating unknown fields is blocked" in audit_record["failure"]


@pytest.mark.asyncio
async def test_update_cell_primary_key_is_blocked_and_failure_is_audited(sample_env):
    service = _service(sample_env)

    outcome = await service.write(
        ops=[WorkbenchSourceWriteOp(op="update_cell", table="Hero", row_id=1, field="ID", new_value=9)],
        reason="change primary key",
    )

    assert outcome.ok is False
    assert outcome.status_code == 400
    assert "Changing the primary key is blocked" in outcome.payload["message"]
    assert outcome.payload["audit_recorded"] is True
    audit_record = _read_audit_lines(service)[0]
    assert audit_record["success"] is False
    assert "Changing the primary key is blocked" in audit_record["failure"]


@pytest.mark.asyncio
async def test_insert_row_requires_object_new_value(sample_env):
    service = _service(sample_env)

    outcome = await service.write(
        ops=[WorkbenchSourceWriteOp(op="insert_row", table="Hero", row_id=3, new_value=["bad"])],
        reason="bad row payload",
    )

    assert outcome.ok is False
    assert outcome.status_code == 400
    assert "insert_row requires new_value to be an object" in outcome.payload["message"]
    assert outcome.payload["audit_recorded"] is True
    audit_record = _read_audit_lines(service)[0]
    assert audit_record["success"] is False
    assert "insert_row requires new_value to be an object" in audit_record["failure"]


@pytest.mark.asyncio
async def test_insert_row_primary_key_mismatch_is_blocked_and_failure_is_audited(sample_env):
    service = _service(sample_env)

    outcome = await service.write(
        ops=[WorkbenchSourceWriteOp(op="insert_row", table="Hero", row_id=3, new_value={"ID": 4, "Name": "Archer", "HP": 90})],
        reason="row id mismatch",
    )

    assert outcome.ok is False
    assert outcome.status_code == 400
    assert "Changing the primary key is blocked" in outcome.payload["message"]
    assert outcome.payload["audit_recorded"] is True
    audit_record = _read_audit_lines(service)[0]
    assert audit_record["success"] is False
    assert "Changing the primary key is blocked" in audit_record["failure"]


@pytest.mark.asyncio
async def test_unknown_op_is_blocked_and_failure_is_audited(sample_env):
    service = _service(sample_env)

    outcome = await service.write(
        ops=[WorkbenchSourceWriteOp(op="explode_world", table="Hero", row_id=1, field="HP", new_value=0)],
        reason="unknown op",
    )

    assert outcome.ok is False
    assert outcome.status_code == 400
    assert "Schema and unsupported ops are blocked" in outcome.payload["message"]
    assert outcome.payload["audit_recorded"] is True
    audit_record = _read_audit_lines(service)[0]
    assert audit_record["success"] is False
    assert "explode_world" in audit_record["failure"]


@pytest.mark.asyncio
async def test_source_write_does_not_change_current_release(sample_env):
    service = _service(sample_env)
    manifest = KnowledgeManifest(
        schema_version="knowledge-manifest.v1",
        release_id="release-h4-001",
        created_at="2026-05-14T00:00:00Z",
        project_root_hash="sha256:project-root",
        source_snapshot_hash="sha256:source-snapshot",
        map_hash="sha256:map-h4-001",
        indexes={
            name: KnowledgeIndexArtifact(path=path, hash="sha256:index", count=0)
            for name, path in DEFAULT_RELEASE_INDEXES.items()
        },
    )
    knowledge_map = KnowledgeMap(
        schema_version="knowledge-map.v1",
        release_id="release-h4-001",
        relationships=[],
    )
    create_release(sample_env["svn_root"], manifest, knowledge_map, indexes={})
    set_current_release(sample_env["svn_root"], "release-h4-001")

    before = get_current_release(sample_env["svn_root"]).release_id

    outcome = await service.write(
        ops=[WorkbenchSourceWriteOp(op="update_cell", table="Hero", row_id=1, field="HP", new_value=150)],
        reason="raise hp",
    )

    after = get_current_release(sample_env["svn_root"]).release_id
    assert outcome.ok is True
    assert outcome.payload["release_id_at_write"] == "release-h4-001"
    assert before == "release-h4-001"
    assert after == "release-h4-001"


def test_append_audit_record_returns_false_when_parent_mkdir_fails(sample_env, monkeypatch):
    service = _service(sample_env)

    def _mkdir(*args, **kwargs):
        raise OSError("mkdir failed")

    monkeypatch.setattr(Path, "mkdir", _mkdir)

    assert service._append_audit_record({"event_type": "workbench.source.write"}) is False


def test_append_audit_record_returns_false_when_open_fails(sample_env, monkeypatch):
    service = _service(sample_env)

    def _open(*args, **kwargs):
        raise OSError("open failed")

    monkeypatch.setattr(Path, "open", _open)

    assert service._append_audit_record({"event_type": "workbench.source.write"}) is False


def test_append_audit_record_returns_false_when_json_serialization_fails(sample_env):
    service = _service(sample_env)

    assert service._append_audit_record({"bad": _JsonBoom()}) is False


@pytest.mark.asyncio
async def test_apply_success_but_audit_failure_returns_write_applied_true(sample_env, monkeypatch):
    service = _service(sample_env)

    monkeypatch.setattr(
        service,
        "_base_audit_payload",
        lambda reason: {
            "event_type": "workbench.source.write",
            "agent_id": "agent-1",
            "session_id": "session-1",
            "time": "2026-05-13T00:00:00+00:00",
            "release_id_at_write": None,
            "reason": reason,
            "bad": _JsonBoom(),
        },
    )

    outcome = await service.write(
        ops=[WorkbenchSourceWriteOp(op="update_cell", table="Hero", row_id=1, field="HP", new_value=150)],
        reason="raise hp",
    )

    assert outcome.ok is False
    assert outcome.status_code == 500
    assert outcome.payload["message"] == "source write applied but audit failed"
    assert outcome.payload["audit_recorded"] is False
    assert outcome.payload["write_applied"] is True
    assert outcome.payload["source_files"] == ["tables/Hero.csv"]
    assert outcome.payload["changes"][0]["new_value"] == 150
    assert _read_csv(sample_env["hero_csv"])[1] == ["1", "Knight", "150"]
    assert not service._audit_path().exists()


@pytest.mark.asyncio
async def test_business_failure_with_failure_audit_failure_returns_original_error(sample_env, monkeypatch):
    service = _service(sample_env)

    monkeypatch.setattr(
        service,
        "_base_audit_payload",
        lambda reason: {
            "event_type": "workbench.source.write",
            "agent_id": "agent-1",
            "session_id": "session-1",
            "time": "2026-05-13T00:00:00+00:00",
            "release_id_at_write": None,
            "reason": reason,
            "bad": _JsonBoom(),
        },
    )

    outcome = await service.write(
        ops=[WorkbenchSourceWriteOp(op="delete_row", table="Hero", row_id=1)],
        reason="forbidden delete",
    )

    assert outcome.ok is False
    assert outcome.status_code == 400
    assert "delete_row is blocked" in outcome.payload["message"]
    assert outcome.payload["audit_recorded"] is False
    assert outcome.payload["write_applied"] is False
    assert not service._audit_path().exists()
