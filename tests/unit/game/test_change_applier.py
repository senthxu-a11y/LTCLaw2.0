from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from ltclaw_gy_x.game.change_applier import ApplyError, ChangeApplier
from ltclaw_gy_x.game.change_proposal import ChangeOp, ChangeProposal
from ltclaw_gy_x.game.config import (
    FilterConfig,
    PathRule,
    ProjectConfig,
    ProjectMeta,
    SvnConfig,
    TableConvention,
)
from ltclaw_gy_x.game.models import TableIndex
from ltclaw_gy_x.game.paths import get_project_raw_table_index_path


def _project_config(svn_root: Path) -> ProjectConfig:
    return ProjectConfig(
        project=ProjectMeta(name="Test Game", engine="Unity", language="zh-CN"),
        svn=SvnConfig(root=str(svn_root), poll_interval_seconds=300, jitter_seconds=30),
        paths=[PathRule(path="tables/*", semantic="table", system="core")],
        filters=FilterConfig(include_ext=[".xlsx", ".csv"], exclude_glob=[]),
        table_convention=TableConvention(header_row=1, primary_key_field="ID"),
        doc_templates={},
        models={},
    )


def _write_csv(path: Path, rows: list[list[object]], delimiter: str = ",") -> None:
    content = "\n".join(delimiter.join("" if value is None else str(value) for value in row) for row in rows)
    path.write_text(content + "\n", encoding="utf-8")


def _read_csv(path: Path, delimiter: str = ",") -> list[list[str]]:
    return [line.split(delimiter) for line in path.read_text(encoding="utf-8").strip().splitlines()]


def _write_xlsx(path: Path, rows: list[list[object]]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    wb.save(path)


def _read_xlsx(path: Path) -> list[list[object]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    return [list(row) for row in ws.iter_rows(values_only=True)]


@pytest.fixture
def sample_env(tmp_path):
    svn_root = tmp_path / "svn"
    tables_dir = svn_root / "tables"
    index_dir = svn_root / ".ltclaw_index" / "tables"
    tables_dir.mkdir(parents=True)
    index_dir.mkdir(parents=True)

    hero_csv = tables_dir / "Hero.csv"
    item_xlsx = tables_dir / "Item.xlsx"
    _write_csv(hero_csv, [["ID", "Name", "HP"], [1, "Knight", 100], [2, "Mage", 80]])
    _write_xlsx(item_xlsx, [["ID", "Name", "Price"], [10, "Potion", 5], [20, "Ether", 12]])

    (index_dir / "Hero.json").write_text(
        '{"schema_version":"table-index.v1","table_name":"Hero","source_path":"tables/Hero.csv","source_hash":"sha256:x","svn_revision":1,"system":"core","row_count":2,"header_row":1,"primary_key":"ID","ai_summary":"","ai_summary_confidence":0.1,"fields":[],"id_ranges":[],"last_indexed_at":"2026-04-28T00:00:00","indexer_model":"test"}',
        encoding="utf-8",
    )
    (index_dir / "Item.json").write_text(
        '{"schema_version":"table-index.v1","table_name":"Item","source_path":"tables/Item.xlsx","source_hash":"sha256:x","svn_revision":1,"system":"core","row_count":2,"header_row":1,"primary_key":"ID","ai_summary":"","ai_summary_confidence":0.1,"fields":[],"id_ranges":[],"last_indexed_at":"2026-04-28T00:00:00","indexer_model":"test"}',
        encoding="utf-8",
    )

    return {
        "svn_root": svn_root,
        "hero_csv": hero_csv,
        "item_xlsx": item_xlsx,
        "project": _project_config(svn_root),
    }


def _proposal(*ops: ChangeOp) -> ChangeProposal:
    return ChangeProposal(id="p1", title="change", ops=list(ops))


@pytest.mark.asyncio
async def test_update_cell_success_for_csv(sample_env):
    applier = ChangeApplier(sample_env["project"], sample_env["svn_root"])

    result = await applier.apply(
        _proposal(ChangeOp(op="update_cell", table="Hero", row_id=1, field="HP", new_value=150))
    )

    assert result["changed_files"] == ["tables/Hero.csv"]
    assert result["summary"] == "1 updates / 0 inserts / 0 deletes"
    assert _read_csv(sample_env["hero_csv"])[1] == ["1", "Knight", "150"]


@pytest.mark.asyncio
async def test_update_cell_success_for_xlsx(sample_env):
    applier = ChangeApplier(sample_env["project"], sample_env["svn_root"])

    result = await applier.apply(
        _proposal(ChangeOp(op="update_cell", table="Item", row_id=10, field="Price", new_value=8))
    )

    assert result["changed_files"] == ["tables/Item.xlsx"]
    assert _read_xlsx(sample_env["item_xlsx"])[1] == [10, "Potion", 8]


@pytest.mark.asyncio
async def test_insert_row_and_delete_row(sample_env):
    applier = ChangeApplier(sample_env["project"], sample_env["svn_root"])

    result = await applier.apply(
        _proposal(
            ChangeOp(
                op="insert_row",
                table="Hero",
                row_id=3,
                new_value={"ID": 3, "Name": "Archer", "HP": 90},
            ),
            ChangeOp(op="delete_row", table="Hero", row_id=1),
        )
    )

    rows = _read_csv(sample_env["hero_csv"])
    assert result["summary"] == "0 updates / 1 inserts / 1 deletes"
    assert rows[1] == ["2", "Mage", "80"]
    assert rows[2] == ["3", "Archer", "90"]


@pytest.mark.asyncio
async def test_dry_run_returns_before_after_preview(sample_env):
    applier = ChangeApplier(sample_env["project"], sample_env["svn_root"])

    preview = await applier.dry_run(
        _proposal(ChangeOp(op="update_cell", table="Hero", row_id=2, field="HP", new_value=95))
    )

    assert preview[0]["ok"] is True
    assert preview[0]["before"] == "80"
    assert preview[0]["after"] == 95
    assert _read_csv(sample_env["hero_csv"])[2] == ["2", "Mage", "80"]


@pytest.mark.asyncio
async def test_missing_row_raises_apply_error(sample_env):
    applier = ChangeApplier(sample_env["project"], sample_env["svn_root"])

    with pytest.raises(ApplyError, match="row_id not found"):
        await applier.apply(
            _proposal(ChangeOp(op="update_cell", table="Hero", row_id=99, field="HP", new_value=1))
        )


@pytest.mark.asyncio
async def test_missing_field_raises_apply_error(sample_env):
    applier = ChangeApplier(sample_env["project"], sample_env["svn_root"])

    with pytest.raises(ApplyError, match="field not found"):
        await applier.apply(
            _proposal(ChangeOp(op="update_cell", table="Hero", row_id=1, field="MP", new_value=1))
        )


@pytest.mark.asyncio
async def test_partial_failure_keeps_written_files_and_cleans_pending(sample_env):
    applier = ChangeApplier(sample_env["project"], sample_env["svn_root"])

    with pytest.raises(ApplyError, match="row_id not found"):
        await applier.apply(
            _proposal(
                ChangeOp(op="update_cell", table="Hero", row_id=1, field="HP", new_value=180),
                ChangeOp(op="update_cell", table="Item", row_id=999, field="Price", new_value=9),
            )
        )

    assert _read_csv(sample_env["hero_csv"])[1] == ["1", "Knight", "180"]
    assert not (sample_env["hero_csv"].with_name("Hero.csv.ltclaw_pending")).exists()
    assert not (sample_env["item_xlsx"].with_name("Item.xlsx.ltclaw_pending")).exists()


@pytest.mark.asyncio
async def test_type_conversion_failure_raises_apply_error(sample_env):
    applier = ChangeApplier(sample_env["project"], sample_env["svn_root"])

    with pytest.raises(ApplyError, match="type conversion failed"):
        await applier.apply(
            _proposal(ChangeOp(op="update_cell", table="Hero", row_id=1, field="HP", new_value="abc"))
        )


def test_read_rows_falls_back_to_project_raw_index(tmp_path):
    svn_root = tmp_path / "svn"
    tables_dir = svn_root / "Tables"
    tables_dir.mkdir(parents=True)
    hero_csv = tables_dir / "HeroTable.csv"
    _write_csv(hero_csv, [["ID", "Name", "HP", "Attack"], [1, "HeroA", 100, 20]])

    raw_index_path = get_project_raw_table_index_path(svn_root, "HeroTable")
    raw_index_path.parent.mkdir(parents=True, exist_ok=True)
    raw_index_path.write_text(
        TableIndex(
            table_name="HeroTable",
            source_path="Tables/HeroTable.csv",
            source_hash="sha256:hero",
            svn_revision=1,
            system="core",
            row_count=1,
            header_row=1,
            primary_key="ID",
            ai_summary="Hero table",
            ai_summary_confidence=0.1,
            fields=[],
            id_ranges=[],
            last_indexed_at="2026-05-15T00:00:00Z",
            indexer_model="rule_only",
        ).model_dump_json(indent=2),
        encoding="utf-8",
    )

    project = ProjectConfig(
        project=ProjectMeta(name="Test Game", engine="Unity", language="zh-CN"),
        svn=SvnConfig(root=str(svn_root), poll_interval_seconds=300, jitter_seconds=30),
        paths=[PathRule(path="Tables/*", semantic="table", system="core")],
        filters=FilterConfig(include_ext=[".csv"], exclude_glob=[]),
        table_convention=TableConvention(header_row=1, primary_key_field="ID"),
        doc_templates={},
        models={},
    )
    applier = ChangeApplier(project, svn_root)

    payload = applier.read_rows("HeroTable", 0, 20)

    assert payload["headers"] == ["ID", "Name", "HP", "Attack"]
    assert payload["rows"] == [["1", "HeroA", "100", "20"]]
    assert payload["source"] == "Tables/HeroTable.csv"

